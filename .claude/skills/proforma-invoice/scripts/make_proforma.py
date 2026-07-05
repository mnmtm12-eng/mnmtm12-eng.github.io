#!/usr/bin/env python3
"""Proforma / commercial invoice generator (xlsx, real export layout).

Example:
    python3 make_proforma.py \
        --buyer "ALPHA TRADING LLC | Baghdad, Iraq | +964 770 000 0000" \
        --item "Armored steel door 90x205 cm:7308.30:100:pcs:120" --deposit 30

--item  DESCRIPTION:HS:QTY:UNIT:UNIT_PRICE   (repeatable)
--buyer lines separated by '|'
Output: <repo>/invoices/PI-YYYYMMDD-HHMM.xlsx
"""
import argparse, json, os
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HERE = os.path.dirname(os.path.abspath(__file__))
EXPORTER = os.path.join(HERE, "..", "data", "exporter_defaults.json")
REPO = os.path.abspath(os.path.join(HERE, "..", "..", "..", ".."))

NAVY, GREY = "1F4E79", "F2F2F2"
THIN = Border(*[Side(style="thin")] * 4)


def parse_item(raw):
    p = raw.rsplit(":", 4)
    if len(p) != 5:
        raise SystemExit(f"bad --item '{raw}' (want DESC:HS:QTY:UNIT:PRICE)")
    return {"desc": p[0], "hs": p[1], "qty": float(p[2]), "unit": p[3], "price": float(p[4])}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--buyer", required=True, help="lines separated by |")
    ap.add_argument("--item", action="append", required=True)
    ap.add_argument("--deposit", type=float, default=30)
    ap.add_argument("--incoterm", default="EXW Kayseri, Turkiye")
    ap.add_argument("--currency", default="USD")
    ap.add_argument("--balance-terms", default="Balance before shipment by T/T")
    ap.add_argument("--exporter", default=EXPORTER)
    args = ap.parse_args()

    exp = json.load(open(args.exporter, encoding="utf-8"))
    items = [parse_item(i) for i in args.item]
    total = sum(i["qty"] * i["price"] for i in items)
    dep_amt = total * args.deposit / 100

    wb = Workbook()
    ws = wb.active
    ws.title = "Proforma"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFG", (5, 42, 11, 9, 8, 13, 14)):
        ws.column_dimensions[col].width = w

    def put(r, c, v, bold=False, size=10, fill=None, color=None, align="left",
            border=False, numfmt=None):
        cell = ws.cell(r, c, v)
        cell.font = Font(bold=bold, size=size, color=color or "000000")
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if fill:
            cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        if border:
            cell.border = THIN

        if numfmt:
            cell.number_format = numfmt

    now = datetime.now()
    pi_no = f"PI-{now:%Y%m%d-%H%M}"
    ws.merge_cells("A1:G1")
    put(1, 1, "PROFORMA INVOICE", bold=True, size=16, fill=NAVY, color="FFFFFF", align="center")

    put(3, 1, "SELLER / EXPORTER", bold=True, fill=GREY)
    put(3, 5, "BUYER / CONSIGNEE", bold=True, fill=GREY)
    seller_lines = [exp["company"], exp["address"], exp["phone"], exp.get("email", ""),
                    f"VKN: {exp.get('tax_office_vkn', '')}"]
    buyer_lines = [b.strip() for b in args.buyer.split("|")]
    for i in range(max(len(seller_lines), len(buyer_lines))):
        r = 4 + i
        if i < len(seller_lines):
            ws.merge_cells(f"A{r}:C{r}")
            put(r, 1, seller_lines[i], size=9)
        if i < len(buyer_lines):
            ws.merge_cells(f"E{r}:G{r}")
            put(r, 5, buyer_lines[i], size=9)

    r = 4 + max(len(seller_lines), len(buyer_lines)) + 1
    put(r, 1, "PI No:", bold=True); put(r, 2, pi_no)
    put(r, 5, "Date:", bold=True); put(r, 6, str(date.today()))
    r += 1
    put(r, 1, "Incoterm:", bold=True); put(r, 2, args.incoterm)
    put(r, 5, "Currency:", bold=True); put(r, 6, args.currency)

    r += 2
    for i, h in enumerate(["#", "Description of goods", "HS Code", "Qty", "Unit",
                           f"Unit price ({args.currency})", f"Amount ({args.currency})"], 1):
        put(r, i, h, bold=True, fill=NAVY, color="FFFFFF", align="center", border=True)
    for n, it in enumerate(items, 1):
        r += 1
        amount = it["qty"] * it["price"]
        row = [n, it["desc"], it["hs"], it["qty"], it["unit"], it["price"], amount]
        for i, v in enumerate(row, 1):
            put(r, i, v, align="center" if i != 2 else "left", border=True,
                numfmt="#,##0.00" if i in (6, 7) else None)
    r += 1
    ws.merge_cells(f"A{r}:E{r}")
    put(r, 1, "TOTAL", bold=True, fill=GREY, align="right", border=True)
    put(r, 6, "", fill=GREY, border=True)
    put(r, 7, total, bold=True, fill=GREY, align="center", border=True, numfmt="#,##0.00")

    r += 2
    put(r, 1, "PAYMENT TERMS", bold=True, fill=GREY)
    r += 1
    ws.merge_cells(f"A{r}:G{r}")
    put(r, 1, f"{args.deposit:.0f}% advance deposit = {dep_amt:,.2f} {args.currency} "
              f"with order. {args.balance_terms}. Goods remain property of seller until full payment.", size=9)

    r += 2
    put(r, 1, "BANK DETAILS", bold=True, fill=GREY)
    b = exp.get("bank", {})
    for label, key in (("Beneficiary", "beneficiary"), ("Bank", "bank_name"),
                       ("Branch", "branch"), ("SWIFT", "swift"),
                       ("IBAN (USD)", "iban_usd"), ("IBAN (EUR)", "iban_eur")):
        if b.get(key):
            r += 1
            put(r, 1, f"{label}:", bold=True, size=9)
            ws.merge_cells(f"B{r}:D{r}")
            put(r, 2, b[key], size=9)

    r += 3
    put(r, 5, "Stamp & signature:", bold=True, size=9)
    ws.merge_cells(f"F{r}:G{r}")
    put(r, 6, "____________________")

    out_dir = os.path.join(REPO, "invoices")
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"{pi_no}.xlsx")
    wb.save(path)
    print(f"OK {path} | total {total:,.2f} {args.currency} | "
          f"deposit {args.deposit:.0f}% = {dep_amt:,.2f}")


if __name__ == "__main__":
    main()
