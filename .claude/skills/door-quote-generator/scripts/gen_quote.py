#!/usr/bin/env python3
"""MN STEEL DOOR - per-country quote / spec-sheet generator (xlsx).

Example:
    python3 gen_quote.py --country egypt --client "Cairo Doors Co." \
        --item 90x205:100:115 --item 120x215:60:135

--item WIDTHxHEIGHT:QTY[:UNIT_PRICE_USD]  (cm; omit price -> TBD)
Output: <repo>/quotes/QT-YYYYMMDD-HHMM-<country>.xlsx
"""
import argparse, json, math, os
from datetime import date, datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HERE = os.path.dirname(os.path.abspath(__file__))
SPECS_FILE = os.path.join(HERE, "..", "data", "country_specs.json")
REPO = os.path.abspath(os.path.join(HERE, "..", "..", "..", ".."))

NAVY, GOLD, GREY = "1F4E79", "D4AF37", "F2F2F2"
THIN = Border(*[Side(style="thin")] * 4)


def parse_item(raw):
    parts = raw.split(":")
    if len(parts) < 2:
        raise SystemExit(f"bad --item '{raw}' (want WxH:QTY[:PRICE])")
    return parts[0], int(parts[1]), (float(parts[2]) if len(parts) > 2 else None)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--country", required=True)
    ap.add_argument("--client", required=True)
    ap.add_argument("--item", action="append", required=True,
                    help="WIDTHxHEIGHT:QTY[:UNIT_PRICE_USD], repeatable")
    ap.add_argument("--out", default=os.path.join(REPO, "quotes"))
    args = ap.parse_args()

    cfg = json.load(open(SPECS_FILE, encoding="utf-8"))
    d = cfg["defaults"]
    ckey = args.country.lower()
    if ckey not in cfg["countries"]:
        raise SystemExit(f"unknown country '{args.country}' — choices: {', '.join(cfg['countries'])}")
    c = cfg["countries"][ckey]
    items = [parse_item(i) for i in args.item]

    wb = Workbook()
    ws = wb.active
    ws.title = "Quote"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", (6, 34, 12, 16, 16)):
        ws.column_dimensions[col].width = w

    def put(row, col, val, bold=False, size=11, fill=None, color=None,
            align="right", border=False, numfmt=None):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(bold=bold, size=size, color=color or "000000")
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if fill:
            cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        if border:
            cell.border = THIN
        if numfmt:
            cell.number_format = numfmt
        return cell

    r = 1
    ws.merge_cells(f"A{r}:E{r}")
    put(r, 1, d["brand"], bold=True, size=18, color=NAVY, align="center")
    r += 1
    ws.merge_cells(f"A{r}:E{r}")
    put(r, 1, d["contact"], size=9, align="center")

    r += 2
    today = date.today()
    qno = f"QT-{today:%Y%m%d}-{ckey.upper()[:3]}"
    ws.merge_cells(f"A{r}:E{r}")
    put(r, 1, "QUOTATION / PROFORMA", bold=True, size=13, fill=NAVY, color="FFFFFF", align="center")
    r += 1
    put(r, 2, "Quote No:", bold=True); put(r, 3, qno, align="left")
    put(r, 4, "Date:", bold=True); put(r, 5, str(today), align="left")
    r += 1
    put(r, 2, "Client:", bold=True); put(r, 3, args.client, align="left")
    put(r, 4, "Valid until:", bold=True)
    put(r, 5, str(today + timedelta(days=d["validity_days"])), align="left")
    r += 1
    put(r, 2, "Market:", bold=True)
    put(r, 3, f"{ckey.title()} ({c['name_ar']})", align="left")

    r += 2
    ws.merge_cells(f"A{r}:E{r}")
    put(r, 1, "SPECIFICATION", bold=True, fill=GOLD, align="center")
    for line in cfg["base_spec"] + c.get("extra_spec", []):
        r += 1
        ws.merge_cells(f"A{r}:E{r}")
        put(r, 1, f"• {line}", size=9, align="left")

    r += 2
    hdr = ["#", "Door size (cm)", "Qty (pcs)", "Unit price (USD)", "Amount (USD)"]
    for i, h in enumerate(hdr, 1):
        put(r, i, h, bold=True, fill=NAVY, color="FFFFFF", align="center", border=True)
    total_qty, total_amt, has_tbd = 0, 0.0, False
    for n, (size, qty, price) in enumerate(items, 1):
        r += 1
        put(r, 1, n, align="center", border=True)
        put(r, 2, size, align="center", border=True)
        put(r, 3, qty, align="center", border=True)
        total_qty += qty
        if price is None:
            has_tbd = True
            put(r, 4, "TBD", align="center", border=True)
            put(r, 5, "TBD", align="center", border=True)
        else:
            put(r, 4, price, align="center", border=True, numfmt="#,##0.00")
            put(r, 5, price * qty, align="center", border=True, numfmt="#,##0.00")
            total_amt += price * qty

    r += 1
    put(r, 2, "TOTAL", bold=True, fill=GREY, border=True)
    put(r, 3, total_qty, bold=True, align="center", fill=GREY, border=True)
    put(r, 4, "", fill=GREY, border=True)
    put(r, 5, ("TBD" if has_tbd and total_amt == 0 else total_amt),
        bold=True, align="center", fill=GREY, border=True, numfmt="#,##0.00")

    dpc = c["doors_per_container"]
    containers = math.ceil(total_qty / dpc)
    r += 2
    ws.merge_cells(f"A{r}:E{r}")
    put(r, 1, f"Loading: {total_qty} doors ≈ {containers} x {d['container']} "
              f"({dpc} doors/container for {ckey.title()})", bold=True, align="left")

    r += 2
    ws.merge_cells(f"A{r}:E{r}")
    put(r, 1, "TERMS", bold=True, fill=GOLD, align="center")
    for t in (f"Delivery: {d['incoterm']}",
              f"Payment: {d['deposit_pct']}% deposit with order, balance before shipment",
              f"Lead time: {d['lead_time_days']} days after deposit",
              "Every container personally quality-inspected before loading."):
        r += 1
        ws.merge_cells(f"A{r}:E{r}")
        put(r, 1, f"• {t}", size=9, align="left")

    os.makedirs(args.out, exist_ok=True)
    path = os.path.join(args.out, f"QT-{today:%Y%m%d}-{datetime.now():%H%M}-{ckey}.xlsx")
    wb.save(path)
    print(f"OK {path} | {total_qty} doors | {containers}x{d['container']} | "
          f"total {'TBD' if has_tbd and total_amt == 0 else f'${total_amt:,.2f}'}")


if __name__ == "__main__":
    main()
