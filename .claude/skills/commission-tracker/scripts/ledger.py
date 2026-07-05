#!/usr/bin/env python3
"""Commission ledger: $1000/container, earned on deposit.

  ledger.py add --client X --containers 2 [--country jordan] [--date YYYY-MM-DD]
                [--rate 1000] [--status pending|deposit|paid] [--note ...]
  ledger.py set-status --id 3 --status deposit
  ledger.py report      # rebuild commission_ledger.xlsx + print totals
"""
import argparse, json, os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "..", "data", "ledger.json")
REPO = os.path.abspath(os.path.join(HERE, "..", "..", "..", ".."))
XLSX = os.path.join(REPO, "commission_ledger.xlsx")

STATUSES = ("pending", "deposit", "paid")
FILLS = {"pending": "FFF2CC", "deposit": "D9EAD3", "paid": "C6E0B4"}
THIN = Border(*[Side(style="thin")] * 4)


def load():
    if os.path.exists(DATA):
        with open(DATA, encoding="utf-8") as f:
            return json.load(f)
    return []


def save(rows):
    os.makedirs(os.path.dirname(DATA), exist_ok=True)
    with open(DATA, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)


def report(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Commissions"
    ws.sheet_view.showGridLines = False
    widths = (5, 12, 24, 12, 11, 10, 13, 10, 28)
    for col, w in zip("ABCDEFGHI", widths):
        ws.column_dimensions[col].width = w

    hdr = ["ID", "Date", "Client", "Country", "Containers", "Rate $", "Commission $", "Status", "Note"]
    for i, h in enumerate(hdr, 1):
        c = ws.cell(1, i, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        c.alignment = Alignment(horizontal="center")
        c.border = THIN

    earned = collected = pend = 0.0
    r = 1
    for e in rows:
        r += 1
        comm = e["containers"] * e["rate"]
        if e["status"] in ("deposit", "paid"):
            earned += comm
        if e["status"] == "paid":
            collected += comm
        if e["status"] == "pending":
            pend += comm
        vals = [e["id"], e["date"], e["client"], e.get("country", ""),
                e["containers"], e["rate"], comm, e["status"], e.get("note", "")]
        for i, v in enumerate(vals, 1):
            c = ws.cell(r, i, v)
            c.border = THIN
            c.alignment = Alignment(horizontal="center" if i not in (3, 9) else "right")
            c.fill = PatternFill(start_color=FILLS[e["status"]],
                                 end_color=FILLS[e["status"]], fill_type="solid")
        ws.cell(r, 7).number_format = "#,##0"

    r += 2
    total_cont = sum(e["containers"] for e in rows)
    summary = [("Total containers", total_cont),
               ("Earned (deposit received)", earned),
               ("Collected (paid to me)", collected),
               ("Still to collect", earned - collected),
               ("Potential (pending deals)", pend)]
    for label, val in summary:
        ws.cell(r, 3, label).font = Font(bold=True)
        c = ws.cell(r, 7, val)
        c.font = Font(bold=True)
        c.number_format = "#,##0"
        r += 1
    ws.freeze_panes = "A2"
    wb.save(XLSX)
    print(f"OK {XLSX}")
    print(f"containers={total_cont} earned=${earned:,.0f} collected=${collected:,.0f} "
          f"to_collect=${earned - collected:,.0f} pending_potential=${pend:,.0f}")


def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)

    a = sub.add_parser("add")
    a.add_argument("--client", required=True)
    a.add_argument("--containers", type=float, required=True)
    a.add_argument("--country", default="")
    a.add_argument("--date", default=str(date.today()))
    a.add_argument("--rate", type=float, default=1000)
    a.add_argument("--status", choices=STATUSES, default="pending")
    a.add_argument("--note", default="")

    s = sub.add_parser("set-status")
    s.add_argument("--id", type=int, required=True)
    s.add_argument("--status", choices=STATUSES, required=True)

    sub.add_parser("report")
    args = ap.parse_args()

    rows = load()
    if args.cmd == "add":
        entry = {"id": (max((e["id"] for e in rows), default=0) + 1),
                 "date": args.date, "client": args.client, "country": args.country,
                 "containers": args.containers, "rate": args.rate,
                 "status": args.status, "note": args.note}
        rows.append(entry)
        save(rows)
        print(f"added #{entry['id']}: {entry['client']} x{entry['containers']} "
              f"[{entry['status']}] -> ${entry['containers'] * entry['rate']:,.0f}")
    elif args.cmd == "set-status":
        hit = next((e for e in rows if e["id"] == args.id), None)
        if not hit:
            raise SystemExit(f"no entry #{args.id}")
        hit["status"] = args.status
        save(rows)
        print(f"#{args.id} {hit['client']} -> {args.status}")
    report(rows)


if __name__ == "__main__":
    main()
