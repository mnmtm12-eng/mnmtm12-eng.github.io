#!/usr/bin/env python3
"""Shipment tracker: deposit->production->inspection->loading->customs->shipped->arrived->closed.

  shipments.py add --client X --country jordan --containers 2 [--eta YYYY-MM-DD] [--note ...]
  shipments.py stage --id 1 --stage production [--note ...] [--date YYYY-MM-DD]
  shipments.py report
"""
import argparse, json, os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "..", "data", "shipments.json")
REPO = os.path.abspath(os.path.join(HERE, "..", "..", "..", ".."))
XLSX = os.path.join(REPO, "shipments.xlsx")

STAGES = ["deposit", "production", "inspection", "loading",
          "customs", "shipped", "arrived", "closed"]
THIN = Border(*[Side(style="thin")] * 4)


def load():
    if os.path.exists(DATA):
        with open(DATA, encoding="utf-8") as f:
            return json.load(f)
    return []


def save(rows):
    os.makedirs(os.path.dirname(DATA), exist_ok=True)
    with open(DATA, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)


def report(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Shipments"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGH", (5, 22, 12, 11, 13, 12, 12, 45)):
        ws.column_dimensions[col].width = w
    hdr = ["ID", "Client", "Country", "Containers", "Stage", "Since", "ETA", "History"]
    for i, h in enumerate(hdr, 1):
        c = ws.cell(1, i, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        c.alignment = Alignment(horizontal="center")
        c.border = THIN
    r = 1
    for s in rows:
        r += 1
        last = s["history"][-1]
        hist = " | ".join(f"{h['date']} {h['stage']}" + (f" ({h['note']})" if h.get("note") else "")
                          for h in s["history"])
        done = last["stage"] == "closed"
        fill = "D9EAD3" if done else ("FCE5CD" if last["stage"] in ("deposit", "production") else "D0E0F0")
        for i, v in enumerate([s["id"], s["client"], s.get("country", ""), s["containers"],
                               last["stage"].upper(), last["date"], s.get("eta", ""), hist], 1):
            c = ws.cell(r, i, v)
            c.border = THIN
            c.alignment = Alignment(horizontal="center" if i < 8 else "left", wrap_text=(i == 8))
            c.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    ws.freeze_panes = "A2"
    wb.save(XLSX)
    open_s = [s for s in rows if s["history"][-1]["stage"] != "closed"]
    print(f"OK {XLSX} | open shipments: {len(open_s)}")
    for s in open_s:
        last = s["history"][-1]
        print(f"  #{s['id']} {s['client']} ({s.get('country','')}) x{s['containers']} "
              f"-> {last['stage'].upper()} since {last['date']} | ETA {s.get('eta','?')}")


def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)

    a = sub.add_parser("add")
    a.add_argument("--client", required=True)
    a.add_argument("--containers", type=float, required=True)
    a.add_argument("--country", default="")
    a.add_argument("--eta", default="")
    a.add_argument("--note", default="")
    a.add_argument("--date", default=str(date.today()))

    st = sub.add_parser("stage")
    st.add_argument("--id", type=int, required=True)
    st.add_argument("--stage", choices=STAGES, required=True)
    st.add_argument("--note", default="")
    st.add_argument("--date", default=str(date.today()))

    sub.add_parser("report")
    args = ap.parse_args()

    rows = load()
    if args.cmd == "add":
        s = {"id": max((x["id"] for x in rows), default=0) + 1,
             "client": args.client, "country": args.country,
             "containers": args.containers, "eta": args.eta,
             "history": [{"date": args.date, "stage": "deposit", "note": args.note}]}
        rows.append(s)
        save(rows)
        print(f"added shipment #{s['id']} {s['client']} x{s['containers']} (stage: deposit)")
    elif args.cmd == "stage":
        s = next((x for x in rows if x["id"] == args.id), None)
        if not s:
            raise SystemExit(f"no shipment #{args.id}")
        s["history"].append({"date": args.date, "stage": args.stage, "note": args.note})
        save(rows)
        print(f"#{args.id} {s['client']} -> {args.stage}")
    report(rows)


if __name__ == "__main__":
    main()
